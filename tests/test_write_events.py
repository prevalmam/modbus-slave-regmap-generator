import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from modbus_slave_regmap_generator.generators import parser, reg_access, write_event
from modbus_slave_regmap_generator.workbook_loader import load_workbook_data
from modbus_slave_regmap_generator.writer import remove_obsolete_generated_files


HEADERS = [
    "",
    "",
    "Reg_Addr",
    "VarName",
    "Type",
    "ArrayLen",
    "Access",
    "Min",
    "Max",
    "Default",
    "NVM_Offset",
    "WRITE_NOTIFY",
    "BUSY_REJECT",
    "WRITE_CHECK",
    "GROUP_VALIDATE",
]


def _write_workbook(path: Path, rows, headers=HEADERS) -> None:
    workbook = Workbook()
    register_table = workbook.active
    register_table.title = "RegisterTable"
    register_table.append(headers)
    for row in rows:
        register_table.append(["", ""] + list(row))
    register_table.append(["", "EOF"])

    length_defs = workbook.create_sheet("LengthDefs")
    length_defs.append(["", "EOF", "", ""])

    config = workbook.create_sheet("Config")
    for _ in range(4):
        config.append([])
    config.append(["", "", "NVM_SIZE", 256])
    config.append(["", "", "SLAVE_ADDR", 1])
    workbook.save(path)


class WriteEventGenerationTests(unittest.TestCase):
    def _load(self, rows, headers=HEADERS):
        temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(temp_dir.cleanup)
        path = Path(temp_dir.name) / "registers.xlsx"
        _write_workbook(path, rows, headers)
        return load_workbook_data(str(path))

    def test_generates_latched_consume_api_for_all_supported_types(self):
        workbook = self._load(
            [
                [1, "command_a", "uint16_t", 1, "RW", 0, 100, 0, "-", "TRUE", "FALSE", "FALSE", "-"],
                [2, "counter", "uint32_t", 1, "RW", 0, 100, 0, "-", "TRUE", "FALSE", "FALSE", "-"],
                [4, "level", "float", 1, "RW", 0, 100, 0, "-", "TRUE", "FALSE", "FALSE", "-"],
                [6, "values", "uint16_t", 3, "RW", 0, 100, 0, "-", "TRUE", "FALSE", "FALSE", "-"],
                [9, "device_name", "string", 8, "RW", "-", "-", "ABC", "-", "TRUE", "FALSE", "FALSE", "-"],
                [13, "status", "uint16_t", 1, "RW", 0, 100, 0, "-", "FALSE", "FALSE", "FALSE", "-"],
            ]
        )

        files = {item.filename: item.content for item in write_event.generate(workbook)}
        header = files["modbus_reg_write_event_slave.h"]
        source = files["modbus_reg_write_event_slave.c"]

        for name in ("command_a", "counter", "level", "values", "device_name"):
            self.assertIn(f"int consume_{name}_written(void);", header)
            self.assertIn(f"int consume_{name}_written(void)", source)
        self.assertNotIn("consume_status_written", header + source)
        self.assertIn("static uint8_t s_written_bits[1U];", source)
        self.assertIn("case MODBUS_IDX_command_a:", source)
        self.assertIn("case MODBUS_IDX_device_name:", source)
        self.assertIn("s_written_bits[0U] |= 0x01U;", source)
        self.assertIn("s_written_bits[0U] |= 0x10U;", source)
        self.assertIn("s_written_bits[0U] &= (uint8_t)(~mask);", source)

    def test_parser_marks_each_accepted_entry_outside_value_change_check(self):
        workbook = self._load(
            [
                [1, "command_a", "uint16_t", 1, "RW", 0, 100, 0, "-", "TRUE", "FALSE", "FALSE", "-"],
                [2, "command_b", "uint16_t", 1, "RW", 0, 100, 0, "-", "TRUE", "FALSE", "FALSE", "-"],
            ]
        )

        parser_source = parser.generate(workbook)[1].content
        changed_block_end = parser_source.index(
            "        modbus_reg_write_event_mark(table_index);"
        )
        changed_block_start = parser_source.rfind(
            "        if (memcmp(entry->ram_ptr, after_value, entry->size) != 0)",
            0,
            changed_block_end,
        )
        self.assertGreater(changed_block_start, 0)
        block = parser_source[changed_block_start:changed_block_end]

        self.assertIn('#include "modbus_reg_write_event_slave.h"', parser_source)
        self.assertTrue(block.rstrip().endswith("}"))
        self.assertIn(
            "        modbus_reg_write_event_mark(table_index);\n"
            "        expected =",
            parser_source,
        )

    def test_internal_setters_do_not_mark_write_events(self):
        workbook = self._load(
            [
                [1, "command", "uint16_t", 1, "RW", 0, 100, 0, "-", "TRUE", "FALSE", "FALSE", "-"],
            ]
        )

        access_source = reg_access.generate(workbook)[1].content

        self.assertNotIn("modbus_reg_write_event", access_source)
        self.assertNotIn("consume_command_written", access_source)

    def test_no_notify_entries_generate_no_op_marker(self):
        workbook = self._load(
            [
                [1, "status", "uint16_t", 1, "RW", 0, 100, 0, "-", "FALSE", "FALSE", "FALSE", "-"],
            ]
        )

        files = {item.filename: item.content for item in write_event.generate(workbook)}
        header = files["modbus_reg_write_event_slave.h"]
        source = files["modbus_reg_write_event_slave.c"]

        self.assertNotIn("consume_", header + source)
        self.assertIn("(void)table_index;", source)

    def test_bitset_expands_past_eight_notified_entries(self):
        rows = [
            [
                index + 1,
                f"command_{index}",
                "uint16_t",
                1,
                "RW",
                0,
                100,
                0,
                "-",
                "TRUE",
                "FALSE",
                "FALSE",
                "-",
            ]
            for index in range(9)
        ]
        workbook = self._load(rows)

        source = write_event.generate(workbook)[0].content

        self.assertIn("static uint8_t s_written_bits[2U];", source)
        self.assertIn("s_written_bits[1U] |= 0x01U;", source)
        self.assertIn(
            "const int written = (s_written_bits[1U] & mask) != 0U;", source
        )

    def test_rejects_legacy_edge_header(self):
        legacy_headers = HEADERS.copy()
        legacy_headers[11] = "EDGE"

        with self.assertRaisesRegex(
            ValueError, "expected 'WRITE_NOTIFY', got 'EDGE'"
        ):
            self._load(
                [
                    [1, "command", "uint16_t", 1, "RW", 0, 100, 0, "-", "TRUE", "FALSE", "FALSE", "-"],
                ],
                legacy_headers,
            )

    def test_rejects_notify_for_read_only_and_reserved_entries(self):
        with self.assertRaisesRegex(
            ValueError, "WRITE_NOTIFY must be FALSE for a read-only register"
        ):
            self._load(
                [
                    [1, "status", "uint16_t", 1, "RO", 0, 100, 0, "-", "TRUE", "FALSE", "FALSE", "-"],
                ]
            )

        with self.assertRaisesRegex(
            ValueError, "reserved WRITE_NOTIFY must be FALSE"
        ):
            self._load(
                [
                    [2, "reserved_2", "reserved", 1, "-", "-", "-", "-", "-", "TRUE", "FALSE", "FALSE", "-"],
                ]
            )

    def test_removes_only_named_obsolete_generated_files(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            obsolete_c = root / "modbus_reg_edge_slave.c"
            obsolete_h = root / "modbus_reg_edge_slave.h"
            keep = root / "user_file.c"
            obsolete_c.write_text("old", encoding="utf-8")
            obsolete_h.write_text("old", encoding="utf-8")
            keep.write_text("keep", encoding="utf-8")

            remove_obsolete_generated_files(
                str(root),
                ("modbus_reg_edge_slave.c", "modbus_reg_edge_slave.h"),
            )

            self.assertFalse(obsolete_c.exists())
            self.assertFalse(obsolete_h.exists())
            self.assertTrue(keep.exists())


if __name__ == "__main__":
    unittest.main()
