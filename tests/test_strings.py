import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from modbus_slave_regmap_generator.generators import (
    parser,
    reg_access,
    reg_map,
    write_guard,
)
from modbus_slave_regmap_generator.workbook_loader import load_workbook_data


HEADERS = [
    "",
    "",
    "Reg_Addr",
    "VarName",
    "Type",
    "ArrayLen",
    "Access",
    "Min",
    "Max",
    "Default",
    "NVM_Offset",
    "UPDATE_NOTIFY",
    "BUSY_REJECT",
    "WRITE_CHECK",
    "GROUP_VALIDATE",
]


def _write_workbook(path: Path, rows) -> None:
    workbook = Workbook()
    register_table = workbook.active
    register_table.title = "RegisterTable"
    register_table.append(HEADERS)
    for row in rows:
        register_table.append(["", ""] + list(row))
    register_table.append(["", "EOF"])

    length_defs = workbook.create_sheet("LengthDefs")
    length_defs.append(["", "EOF", "", ""])

    config = workbook.create_sheet("Config")
    for _ in range(4):
        config.append([])
    config.append(["", "", "NVM_SIZE", 256])
    config.append(["", "", "SLAVE_ADDR", 1])
    workbook.save(path)


class FixedStringGenerationTests(unittest.TestCase):
    def _load(self, rows):
        temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(temp_dir.cleanup)
        path = Path(temp_dir.name) / "registers.xlsx"
        _write_workbook(path, rows)
        return load_workbook_data(str(path))

    def _string_row(self, default="ABCDEFGH", *, group="NAME"):
        return [
            1000,
            "device_name",
            "string",
            8,
            "RW",
            "-",
            "-",
            default,
            "0x20",
            "TRUE",
            "FALSE",
            "TRUE",
            group,
        ]

    def test_array_len_is_field_size_and_maximum_character_count(self):
        workbook = self._load([self._string_row()])
        entry = workbook.entries[0]

        self.assertEqual(8, entry["length"])
        self.assertEqual("sizeof(char) * 8", entry["size"])
        self.assertEqual("ABCDEFGH", entry["default_value"])
        self.assertEqual(
            "static char device_name[8] = "
            "{0x41, 0x42, 0x43, 0x44, 0x45, 0x46, 0x47, 0x48};",
            entry["ram_decl"],
        )

    def test_rejects_default_longer_than_array_len(self):
        with self.assertRaisesRegex(
            ValueError, "string Default length must be 8 bytes or less"
        ):
            self._load([self._string_row("ABCDEFGHI")])

    def test_rejects_odd_array_len(self):
        row = self._string_row("ABC")
        row[3] = 7
        with self.assertRaisesRegex(ValueError, "string ArrayLen must be even"):
            self._load([row])

    def test_supports_explicit_empty_and_literal_dash_defaults(self):
        empty = self._load([self._string_row("-", group="-")]).entries[0]
        literal = self._load([self._string_row('"-"', group="-")]).entries[0]

        self.assertEqual("", empty["default_value"])
        self.assertEqual("-", literal["default_value"])
        self.assertIn("{0x00, 0x00", empty["ram_decl"])
        self.assertIn("{0x2D, 0x00", literal["ram_decl"])

    def test_generates_exact_size_zero_padded_default_fields(self):
        workbook = self._load([self._string_row("ABC")])
        source = reg_map.generate(workbook)[1].content

        expected = (
            "{0x41, 0x42, 0x43, 0x00, 0x00, 0x00, 0x00, 0x00}"
        )
        self.assertIn(f"static char device_name[8] = {expected};", source)
        self.assertIn(f"const char default_device_name[8] = {expected};", source)
        self.assertIn("sizeof(char) * 8,", source)

    def test_generates_shared_bounded_field_helpers(self):
        workbook = self._load([self._string_row()])
        files = {item.filename: item.content for item in reg_map.generate(workbook)}
        header = files["modbus_reg_map_slave.h"]
        source = files["modbus_reg_map_slave.c"]

        self.assertIn("uint16_t modbus_string_field_length(", header)
        self.assertIn("int modbus_string_field_is_valid(", header)
        self.assertIn("while ((length < field_size)", source)
        self.assertNotIn("\x00", source)
        self.assertIn("if (found_nul != 0)", source)
        self.assertIn("return 1;", source)

    def test_accessors_expose_only_safe_copy_getter(self):
        workbook = self._load([self._string_row()])
        files = {item.filename: item.content for item in reg_access.generate(workbook)}
        header = files["modbus_reg_access_slave.h"]
        source = files["modbus_reg_access_slave.c"]

        self.assertNotIn("const char *get_device_name(void);", header)
        self.assertIn("#define MODBUS_device_name_MAX_LENGTH  (8U)", header)
        self.assertIn(
            "#define MODBUS_device_name_BUFFER_SIZE "
            "(MODBUS_device_name_MAX_LENGTH + 1U)",
            header,
        )
        self.assertIn(
            "int get_device_name_copy(char *dst, uint16_t dst_size);", header
        )
        self.assertIn("int set_device_name(const char *value);", header)
        self.assertIn(
            "const uint16_t len = modbus_string_field_length(src, 8U);", source
        )
        self.assertIn("const uint16_t len = string_length_limited(value, 9U);", source)
        self.assertIn("if (is_valid_string_value(value, 8U) == 0)", source)
        self.assertIn(
            "write_nvm_if_used(g_reg_table_slave[MODBUS_IDX_device_name].nvm_offset, "
            "ram, g_reg_table_slave[MODBUS_IDX_device_name].size);",
            source,
        )

    def test_parser_accepts_valid_full_fields_without_hidden_nul(self):
        workbook = self._load([self._string_row()])
        source = parser.generate(workbook)[1].content

        self.assertIn(
            "if (modbus_string_field_is_valid(\n"
            "                    (const char *)data, entry->size) == 0)",
            source,
        )
        self.assertIn("(void)memcpy(after_value, data, entry->size);", source)
        self.assertNotIn("return (found_nul != 0) ? 0 : -1;", source)

    def test_string_write_guards_receive_length_aware_views(self):
        workbook = self._load([self._string_row()])
        header = write_guard.generate(workbook)[0].content
        source = parser.generate(workbook)[1].content

        self.assertIn("} modbus_string_view_t;", header)
        self.assertIn("modbus_string_view_t device_name;", header)
        self.assertIn("const modbus_string_view_t *current_value", header)
        self.assertIn("const modbus_string_view_t *new_value", header)
        self.assertIn("modbus_string_view_t current_value;", source)
        self.assertIn("new_value.length = modbus_string_field_length(", source)
        self.assertIn("snapshot->device_name.length =", source)


if __name__ == "__main__":
    unittest.main()
