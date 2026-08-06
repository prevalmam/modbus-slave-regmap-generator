import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from modbus_slave_regmap_generator.generators import parser, reg_access, update_notify
from modbus_slave_regmap_generator.workbook_loader import load_workbook_data
from modbus_slave_regmap_generator.writer import remove_obsolete_generated_files


HEADERS = [
    "",
    "",
    "Reg_Addr",
    "VarName",
    "Type",
    "ArrayLen",
    "Access",
    "Min",
    "Max",
    "Default",
    "NVM_Offset",
    "UPDATE_NOTIFY",
    "BUSY_REJECT",
    "WRITE_CHECK",
    "GROUP_VALIDATE",
]


def _write_workbook(path: Path, rows, headers=HEADERS) -> None:
    workbook = Workbook()
    register_table = workbook.active
    register_table.title = "RegisterTable"
    register_table.append(headers)
    for row in rows:
        register_table.append(["", ""] + list(row))
    register_table.append(["", "EOF"])

    length_defs = workbook.create_sheet("LengthDefs")
    length_defs.append(["", "EOF", "", ""])

    config = workbook.create_sheet("Config")
    for _ in range(4):
        config.append([])
    config.append(["", "", "NVM_SIZE", 256])
    config.append(["", "", "SLAVE_ADDR", 1])
    workbook.save(path)


class UpdateNotifyGenerationTests(unittest.TestCase):
    def _load(self, rows, headers=HEADERS):
        temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(temp_dir.cleanup)
        path = Path(temp_dir.name) / "registers.xlsx"
        _write_workbook(path, rows, headers)
        return load_workbook_data(str(path))

    def test_generates_required_callbacks_and_source_dispatch_for_supported_types(self):
        workbook = self._load(
            [
                [1, "command_a", "uint16_t", 1, "RW", 0, 100, 0, "-", "TRUE", "FALSE", "FALSE", "-"],
                [2, "counter", "uint32_t", 1, "RW", 0, 100, 0, "-", "TRUE", "FALSE", "FALSE", "-"],
                [4, "level", "float", 1, "RW", 0, 100, 0, "-", "TRUE", "FALSE", "FALSE", "-"],
                [6, "values", "uint16_t", 3, "RW", 0, 100, 0, "-", "TRUE", "FALSE", "FALSE", "-"],
                [9, "device_name", "string", 8, "RW", "-", "-", "ABC", "-", "TRUE", "FALSE", "FALSE", "-"],
                [13, "status", "uint16_t", 1, "RW", 0, 100, 0, "-", "FALSE", "FALSE", "FALSE", "-"],
            ]
        )

        files = {item.filename: item.content for item in update_notify.generate(workbook)}
        header = files["modbus_reg_update_notify_slave.h"]
        source = files["modbus_reg_update_notify_slave.c"]

        self.assertIn("MODBUS_REG_UPDATE_SOURCE_MASTER_WRITE = 1", header)
        self.assertIn("MODBUS_REG_UPDATE_SOURCE_INTERNAL_SET = 2", header)
        for name in ("command_a", "counter", "level", "values", "device_name"):
            self.assertIn(f"extern void modbus_user_{name}_updated(", header)
            self.assertIn(f"modbus_user_{name}_updated(", source)
        self.assertNotIn("modbus_user_status_updated", header + source)
        self.assertNotIn("consume_", header + source)
        self.assertNotIn("s_written_bits", source)
        self.assertIn("case MODBUS_IDX_command_a:", source)
        self.assertIn("case MODBUS_IDX_device_name:", source)

    def test_master_notify_covers_each_full_entry_in_address_order(self):
        workbook = self._load(
            [
                [5, "later", "uint16_t", 1, "RW", 0, 100, 0, "-", "TRUE", "FALSE", "FALSE", "-"],
                [1, "first", "uint32_t", 1, "RW", 0, 100, 0, "-", "TRUE", "FALSE", "FALSE", "-"],
                [3, "middle", "uint16_t", 2, "RW", 0, 100, 0, "-", "TRUE", "FALSE", "FALSE", "-"],
            ]
        )

        source = update_notify.generate(workbook)[0].content

        self.assertLess(
            source.index("modbus_user_first_updated("),
            source.index("modbus_user_middle_updated("),
        )
        self.assertLess(
            source.index("modbus_user_middle_updated("),
            source.index("modbus_user_later_updated("),
        )
        self.assertIn("((uint32_t)start_addr <= 1UL)", source)
        self.assertIn("(end_addr >= 3UL)", source)
        self.assertIn("((uint32_t)start_addr <= 3UL)", source)
        self.assertIn("(end_addr >= 5UL)", source)

    def test_parser_notifies_after_ack_request_on_success_only(self):
        workbook = self._load(
            [
                [1, "command_a", "uint16_t", 1, "RW", 0, 100, 0, "-", "TRUE", "FALSE", "FALSE", "-"],
                [2, "command_b", "uint16_t", 1, "RW", 0, 100, 0, "-", "TRUE", "FALSE", "FALSE", "-"],
            ]
        )

        parser_source = parser.generate(workbook)[1].content

        self.assertIn('#include "modbus_reg_update_notify_slave.h"', parser_source)
        self.assertIn(
            "modbus_send_write_single_ack(slave_addr, start_addr, data);\n"
            "                        modbus_reg_update_notify_master_write(start_addr, 1U);",
            parser_source,
        )
        self.assertIn(
            "modbus_send_write_multi_ack(slave_addr, start_addr, num_regs);\n"
            "                            modbus_reg_update_notify_master_write(start_addr, num_regs);",
            parser_source,
        )
        commit_start = parser_source.index(
            "        if (memcmp(entry->ram_ptr, after_value, entry->size) != 0)"
        )
        commit_end = parser_source.index("    return 0;", commit_start)
        self.assertNotIn("update_notify", parser_source[commit_start:commit_end])

    def test_internal_setters_notify_after_success_including_same_value(self):
        workbook = self._load(
            [
                [1, "command", "uint16_t", 1, "RW", 0, 100, 0, "-", "TRUE", "FALSE", "FALSE", "-"],
                [2, "values", "uint16_t", 2, "RW", 0, 100, 0, "-", "TRUE", "FALSE", "FALSE", "-"],
                [4, "device_name", "string", 8, "RW", "-", "-", "ABC", "-", "TRUE", "FALSE", "FALSE", "-"],
                [8, "plain", "uint16_t", 1, "RW", 0, 100, 0, "-", "FALSE", "FALSE", "FALSE", "-"],
            ]
        )

        source = reg_access.generate(workbook)[1].content

        self.assertIn('#include "modbus_reg_update_notify_slave.h"', source)
        for name in ("command", "values", "device_name"):
            function_start = source.index(f"int set_{name}(")
            function_end = source.index("\n}", function_start)
            function = source[function_start:function_end]
            self.assertIn("if (value != current)" if name != "device_name" else "if (memcmp(ram, temp, sizeof(temp)) != 0)", function)
            self.assertIn(
                f"modbus_reg_update_notify_internal_set(MODBUS_IDX_{name});",
                function,
            )
            self.assertTrue(function.rstrip().endswith("return 1;"))
        plain_start = source.index("int set_plain(")
        plain_end = source.index("\n}", plain_start)
        self.assertNotIn("update_notify", source[plain_start:plain_end])

        masked_start = source.index("int set_command_masked(")
        masked_end = source.index("\n}", masked_start)
        masked_function = source[masked_start:masked_end]
        self.assertIn("return set_command(current);", masked_function)
        self.assertNotIn("update_notify", masked_function)

    def test_read_only_can_notify_internal_set_and_reserved_cannot_notify(self):
        workbook = self._load(
            [
                [1, "status", "uint16_t", 1, "RO", 0, 100, 0, "-", "TRUE", "FALSE", "FALSE", "-"],
            ]
        )
        self.assertTrue(workbook.entries[0]["update_notify"])

        with self.assertRaisesRegex(
            ValueError, "reserved UPDATE_NOTIFY must be FALSE"
        ):
            self._load(
                [
                    [2, "reserved_2", "reserved", 1, "-", "-", "-", "-", "-", "TRUE", "FALSE", "FALSE", "-"],
                ]
            )

    def test_rejects_legacy_write_notify_and_edge_headers(self):
        for legacy_header in ("WRITE_NOTIFY", "EDGE"):
            with self.subTest(legacy_header=legacy_header):
                legacy_headers = HEADERS.copy()
                legacy_headers[11] = legacy_header
                with self.assertRaisesRegex(
                    ValueError,
                    f"expected 'UPDATE_NOTIFY', got '{legacy_header}'",
                ):
                    self._load(
                        [
                            [1, "command", "uint16_t", 1, "RW", 0, 100, 0, "-", "TRUE", "FALSE", "FALSE", "-"],
                        ],
                        legacy_headers,
                    )

    def test_no_notify_entries_generate_no_op_helpers(self):
        workbook = self._load(
            [
                [1, "status", "uint16_t", 1, "RW", 0, 100, 0, "-", "FALSE", "FALSE", "FALSE", "-"],
            ]
        )

        files = {item.filename: item.content for item in update_notify.generate(workbook)}
        header = files["modbus_reg_update_notify_slave.h"]
        source = files["modbus_reg_update_notify_slave.c"]

        self.assertNotIn("extern void modbus_user_", header)
        self.assertIn("(void)start_addr;", source)
        self.assertIn("(void)num_regs;", source)
        self.assertIn("(void)table_index;", source)

    def test_removes_retired_generated_files(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            obsolete_names = (
                "modbus_reg_edge_slave.c",
                "modbus_reg_edge_slave.h",
                "modbus_reg_write_event_slave.c",
                "modbus_reg_write_event_slave.h",
            )
            for name in obsolete_names:
                (root / name).write_text("old", encoding="utf-8")
            keep = root / "user_file.c"
            keep.write_text("keep", encoding="utf-8")

            remove_obsolete_generated_files(str(root), obsolete_names)

            for name in obsolete_names:
                self.assertFalse((root / name).exists())
            self.assertTrue(keep.exists())


if __name__ == "__main__":
    unittest.main()
