import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from modbus_slave_regmap_generator.generators import parser, reg_access, reg_edge
from modbus_slave_regmap_generator.workbook_loader import load_workbook_data


HEADERS = [
    "",
    "",
    "Reg_Addr",
    "VarName",
    "Type",
    "ArrayLen",
    "Access",
    "Min",
    "Max",
    "Default",
    "NVM_Offset",
    "EDGE",
    "BUSY_REJECT",
    "WRITE_CHECK",
    "GROUP_VALIDATE",
]


def _write_workbook(path: Path, rows) -> None:
    workbook = Workbook()
    register_table = workbook.active
    register_table.title = "RegisterTable"
    register_table.append(HEADERS)
    for row in rows:
        register_table.append(["", ""] + list(row))
    register_table.append(["", "EOF"])

    length_defs = workbook.create_sheet("LengthDefs")
    length_defs.append(["", "EOF", "", ""])

    config = workbook.create_sheet("Config")
    for _ in range(4):
        config.append([])
    config.append(["", "", "NVM_SIZE", 256])
    config.append(["", "", "SLAVE_ADDR", 1])
    workbook.save(path)


class RegisterEdgeSyncGenerationTests(unittest.TestCase):
    def _load(self, rows):
        temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(temp_dir.cleanup)
        path = Path(temp_dir.name) / "registers.xlsx"
        _write_workbook(path, rows)
        return load_workbook_data(str(path))

    def test_scalar_setter_syncs_only_its_own_edge_state(self):
        workbook = self._load(
            [
                [1, "command_a", "uint16_t", 1, "RW", 0, 100, 0, "-", "TRUE", "FALSE", "FALSE", "-"],
                [2, "command_b", "uint16_t", 1, "RW", 0, 100, 0, "-", "TRUE", "FALSE", "FALSE", "-"],
                [3, "status", "uint16_t", 1, "RW", 0, 100, 0, "-", "FALSE", "FALSE", "FALSE", "-"],
            ]
        )

        access_source = reg_access.generate(workbook)[1].content
        edge_files = {item.filename: item.content for item in reg_edge.generate(workbook)}
        edge_header = edge_files["modbus_reg_edge_slave.h"]
        edge_source = edge_files["modbus_reg_edge_slave.c"]
        parser_source = parser.generate(workbook)[1].content

        self.assertIn('#include "modbus_reg_edge_slave.h"', access_source)
        self.assertEqual(
            access_source.count("modbus_reg_edge_sync_command_a(value);"), 2
        )
        self.assertEqual(
            access_source.count("modbus_reg_edge_sync_command_b(value);"), 2
        )
        self.assertNotIn("modbus_reg_edge_sync_status", access_source)
        self.assertNotIn("modbus_reg_edge_init();", access_source)
        self.assertNotIn("modbus_reg_edge_sync_", parser_source)

        self.assertIn(
            "void modbus_reg_edge_sync_command_a(uint16_t value);", edge_header
        )
        self.assertIn("static uint16_t s_prev_command_a_rising;", edge_source)
        self.assertIn("static uint16_t s_prev_command_a_falling;", edge_source)
        self.assertIn("static uint16_t s_prev_command_a_toggled;", edge_source)
        self.assertIn("s_prev_command_a_rising = value;", edge_source)
        self.assertIn("s_prev_command_a_falling = value;", edge_source)
        self.assertIn("s_prev_command_a_toggled = value;", edge_source)

    def test_array_setter_syncs_changed_index_and_any_changed_state(self):
        workbook = self._load(
            [
                [1, "values", "uint32_t", 3, "RW", 0, 100, 0, "-", "TRUE", "FALSE", "FALSE", "-"],
                [7, "levels", "float", 2, "RW", 0, 100, 0, "-", "TRUE", "FALSE", "FALSE", "-"],
            ]
        )

        access_source = reg_access.generate(workbook)[1].content
        edge_files = {item.filename: item.content for item in reg_edge.generate(workbook)}
        edge_header = edge_files["modbus_reg_edge_slave.h"]
        edge_source = edge_files["modbus_reg_edge_slave.c"]

        self.assertEqual(
            access_source.count("modbus_reg_edge_sync_values(index, value);"), 2
        )
        self.assertEqual(
            access_source.count("modbus_reg_edge_sync_levels(index, value);"), 2
        )
        self.assertIn(
            "void modbus_reg_edge_sync_values(uint16_t index, uint32_t value);",
            edge_header,
        )
        self.assertIn("s_prev_values_rising[index] = value;", edge_source)
        self.assertIn("s_prev_values_falling[index] = value;", edge_source)
        self.assertIn("s_prev_values_toggled[index] = value;", edge_source)
        self.assertIn("s_prev_values_any_changed[index] = value;", edge_source)
        self.assertIn(
            "void modbus_reg_edge_sync_levels(uint16_t index, float value);",
            edge_header,
        )
        self.assertIn("s_prev_levels_changed[index] = value;", edge_source)
        self.assertIn("s_prev_levels_any_changed[index] = value;", edge_source)

    def test_no_edge_entries_do_not_add_edge_dependency(self):
        workbook = self._load(
            [
                [1, "status", "uint16_t", 1, "RW", 0, 100, 0, "-", "FALSE", "FALSE", "FALSE", "-"],
            ]
        )

        access_source = reg_access.generate(workbook)[1].content

        self.assertNotIn("modbus_reg_edge_slave.h", access_source)
        self.assertNotIn("modbus_reg_edge_sync_", access_source)


if __name__ == "__main__":
    unittest.main()
