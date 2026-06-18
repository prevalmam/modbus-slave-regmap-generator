import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from modbus_slave_regmap_generator.generators import parser, write_guard
from modbus_slave_regmap_generator.workbook_loader import load_workbook_data


HEADERS = [
    "",
    "",
    "Reg_Addr",
    "VarName",
    "Type",
    "ArrayLen",
    "Access",
    "Min",
    "Max",
    "Default",
    "NVM_Offset",
    "EDGE",
    "BUSY_REJECT",
    "WRITE_CHECK",
    "GROUP_VALIDATE",
]


def _write_workbook(path: Path, rows) -> None:
    workbook = Workbook()
    register_table = workbook.active
    register_table.title = "RegisterTable"
    register_table.append(HEADERS)
    for row in rows:
        register_table.append(["", ""] + list(row))
    register_table.append(["", "EOF"])

    length_defs = workbook.create_sheet("LengthDefs")
    length_defs.append(["", "EOF", "", ""])

    config = workbook.create_sheet("Config")
    for _ in range(4):
        config.append([])
    config.append(["", "", "NVM_SIZE", 256])
    config.append(["", "", "SLAVE_ADDR", 1])
    workbook.save(path)


class WriteGuardGenerationTests(unittest.TestCase):
    def _load(self, rows):
        temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(temp_dir.cleanup)
        path = Path(temp_dir.name) / "registers.xlsx"
        _write_workbook(path, rows)
        return load_workbook_data(str(path))

    def test_generates_typed_callbacks_pointer_snapshot_and_busy_api(self):
        workbook = self._load(
            [
                [1, "lower", "uint16_t", 1, "RW", 0, 100, 10, "-", "FALSE", "TRUE", "TRUE", "LOWER_UPPER"],
                [2, "upper", "uint16_t", 1, "RW", 0, 100, 20, "-", "FALSE", "FALSE", "TRUE", "LOWER_UPPER"],
                [3, "table", "uint16_t", 2, "RW", 0, 100, 0, "-", "FALSE", "FALSE", "TRUE", "TABLE"],
                [5, "name", "string", 8, "RW", "-", "-", "ABC", "-", "FALSE", "FALSE", "TRUE", "NAME"],
            ]
        )

        files = {item.filename: item.content for item in write_guard.generate(workbook)}
        header = files["modbus_reg_write_guard_slave.h"]
        source = files["modbus_reg_write_guard_slave.c"]
        parser_source = parser.generate(workbook)[1].content

        self.assertIn("typedef uint8_t MB_BOOL;", header)
        self.assertIn("typedef uint8_t modbus_write_result_t;", header)
        self.assertIn("MODBUS_WRITE_ILLEGAL_VALUE", header)
        self.assertIn("const uint16_t *lower;", header)
        self.assertIn("const uint16_t *table;", header)
        self.assertIn("const char *name;", header)
        self.assertIn("modbus_user_write_check_lower(", header)
        self.assertIn("const uint16_t current_value[]", header)
        self.assertIn("const char current_value[]", header)
        self.assertIn("modbus_user_group_validate_lower_upper(", header)
        self.assertIn(
            "extern modbus_write_result_t modbus_user_write_check_lower(", header
        )
        self.assertIn(
            "extern modbus_write_result_t modbus_user_group_validate_lower_upper(",
            header,
        )
        self.assertIn("modbus_get_busy_reject_lower(void)", header)
        self.assertIn("s_busy_reject_bits[1U]", source)
        self.assertNotIn("BR_", header + source)

        self.assertIn("static modbus_write_scratch_t s_write_scratch;", parser_source)
        self.assertIn("snapshot->lower =", parser_source)
        self.assertIn("group_lower_upper_affected", parser_source)
        self.assertIn("MODBUS_EXCEPTION_SLAVE_DEVICE_BUSY", parser_source)
        self.assertIn("normalize_user_write_result", parser_source)
        self.assertIn("return MODBUS_WRITE_DEVICE_FAILURE;", parser_source)
        self.assertIn("return (int)user_result;", parser_source)
        self.assertNotIn("MODBUS_EXC_SLAVE_DEVICE_BUSY", parser_source)

    def test_rejects_non_boolean_busy_reject(self):
        with self.assertRaisesRegex(ValueError, "BUSY_REJECT must be TRUE or FALSE"):
            self._load(
                [
                    [1, "mode", "uint16_t", 1, "RW", 0, 1, 0, "-", "FALSE", "-", "FALSE", "-"],
                ]
            )

    def test_rejects_lowercase_boolean(self):
        with self.assertRaisesRegex(ValueError, "BUSY_REJECT must be TRUE or FALSE"):
            self._load(
                [
                    [1, "mode", "uint16_t", 1, "RW", 0, 1, 0, "-", "FALSE", "true", "FALSE", "-"],
                ]
            )

    def test_accepts_excel_boolean_cells(self):
        workbook = self._load(
            [
                [1, "mode", "uint16_t", 1, "RW", 0, 1, 0, "-", False, True, False, "-"],
            ]
        )

        entry = workbook.entries[0]
        self.assertFalse(entry["edge"])
        self.assertTrue(entry["busy_reject"])
        self.assertFalse(entry["write_check"])

    def test_rejects_empty_group_validate(self):
        with self.assertRaisesRegex(ValueError, "GROUP_VALIDATE is empty"):
            self._load(
                [
                    [1, "mode", "uint16_t", 1, "RW", 0, 1, 0, "-", "FALSE", "FALSE", "FALSE", ""],
                ]
            )

    def test_rejects_write_guard_on_read_only_register(self):
        with self.assertRaisesRegex(ValueError, "WRITE_CHECK must be FALSE"):
            self._load(
                [
                    [1, "status", "uint16_t", 1, "RO", 0, 1, 0, "-", "FALSE", "FALSE", "TRUE", "-"],
                ]
            )


if __name__ == "__main__":
    unittest.main()
